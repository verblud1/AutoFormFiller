"""Модуль для сортировки ячеек Excel по цветам"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import PatternFill as OpenpyxlPatternFill, GradientFill
from collections import defaultdict


def get_cell_background_color(cell):
    """Получение цвета фона ячейки"""
    # Проверяем fill_type - если None или 'none', значит нет заливки
    if cell.fill.fill_type is None or cell.fill.fill_type == 'none':
        return None
    
    # Получаем цвета из разных возможных источников
    color_attrs = ['fgColor', 'bgColor', 'start_color', 'end_color']
    
    for attr_name in color_attrs:
        if hasattr(cell.fill, attr_name):
            color_obj = getattr(cell.fill, attr_name)
            if color_obj is not None and hasattr(color_obj, 'rgb') and color_obj.rgb is not None:
                rgb_color = color_obj.rgb
                # Проверяем, что это не "пустой" цвет (черный или прозрачный)
                if rgb_color not in ['00000000', '000000']:
                    return rgb_color
    
    return None


def categorize_by_color(workbook, sheet_name=None):
    """Классификация ячеек по цветам"""
    if sheet_name:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active
    
    color_categories = {
        'red': [],      # не ответили на звонок
        'yellow': [],   # занесено в базу данных
        'green': [],    # занесено в базу данных
        'gray': [],     # только позвонили, но не занесли в базу данных
        'none': []      # еще не сделано
    }
    
    # Определение RGB значений для различных цветов
    red_colors = [
        'FFFF0000', 'FFFE0000', 'FFFD0000',  # Чистый красный и близкие оттенки
        'FFCC0000', 'FFC00000', 'FF0000FF',  # Другие варианты красного
        'FF990000', 'FF800000', 'FF660000',
        'FFEE0000', 'FFDD0000', 'FFBB0000'
    ]
    
    yellow_colors = [
        'FFFFFF00', 'FFFFFE00', 'FFFFFD00',  # Чистый желтый и близкие оттенки
        'FFFFFFCC', 'FFFFFF99', 'FFFFFF66',
        'FFFFCC00', 'FFFF9900', 'FFFFCC33',
        'FFFFFF33', 'FFFFEE00', 'FFFFDD00'
    ]
    
    green_colors = [
        'FF00FF00', 'FF00FE00', 'FF00FD00',  # Чистый зеленый и близкие оттенки
        'FF00CC00', 'FF009900', 'FF00CC33',
        'FF33CC33', 'FF00AA00', 'FF008800',
        'FF00EE00', 'FF00DD00', 'FF00BB00'
    ]
    
    gray_colors = [
        'FF808080', 'FF7F7F7F', 'FF7E7E7E',  # Серый и близкие оттенки
        'FFA9A9A9', 'FF888888', 'FF696969',
        'FFC0C0C0', 'FFD3D3D3', 'FFDCDCDC',
        'FFEEEEEE', 'FFDDDDDD', 'FFCCCCCC'
    ]
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            cell_color = get_cell_background_color(cell)
            
            # Если цвет не определен, добавляем в категорию 'none'
            if cell_color is None or cell_color == '00000000' or cell_color == '000000':  # 00000000 или 000000 означает отсутствие цвета
                color_categories['none'].append({
                    'cell': cell.coordinate,
                    'value': cell.value,
                    'row': row_idx,
                    'col': col_idx
                })
            else:
                # Преобразуем цвет к формату без альфа-канала, если он присутствует
                formatted_color = cell_color.upper()
                if len(formatted_color) == 8:
                    # Убираем альфа-канал (первые 2 символа)
                    formatted_color = formatted_color[2:]
                
                # Сравниваем с известными цветами
                if any(color.endswith(formatted_color) or formatted_color.endswith(color[2:]) for color in red_colors):
                    color_categories['red'].append({
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'row': row_idx,
                        'col': col_idx
                    })
                elif any(color.endswith(formatted_color) or formatted_color.endswith(color[2:]) for color in yellow_colors):
                    color_categories['yellow'].append({
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'row': row_idx,
                        'col': col_idx
                    })
                elif any(color.endswith(formatted_color) or formatted_color.endswith(color[2:]) for color in green_colors):
                    color_categories['green'].append({
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'row': row_idx,
                        'col': col_idx
                    })
                elif any(color.endswith(formatted_color) or formatted_color.endswith(color[2:]) for color in gray_colors):
                    color_categories['gray'].append({
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'row': row_idx,
                        'col': col_idx
                    })
                else:
                    # Если цвет не определен как один из основных, добавляем в 'none'
                    color_categories['none'].append({
                        'cell': cell.coordinate,
                        'value': cell.value,
                        'row': row_idx,
                        'col': col_idx
                    })
    
    return color_categories


def sort_excel_by_colors(file_path, output_path=None, sheet_name=None):
    """Сортировка Excel файла по цветам ячеек"""
    # Загружаем рабочую книгу
    wb = openpyxl.load_workbook(file_path)
    
    # Классифицируем ячейки по цветам
    categories = categorize_by_color(wb, sheet_name)
    
    # Если указан путь для вывода, создаем новый файл с результатами
    if output_path:
        # Создаем новую рабочую книгу для результатов
        result_wb = openpyxl.Workbook()
        
        # Удаляем лист по умолчанию
        result_wb.remove(result_wb.active)
        
        # Создаем листы для каждой категории
        for category_name, cells in categories.items():
            if cells:  # Только если есть данные в категории
                ws = result_wb.create_sheet(title=category_name.capitalize())
                
                # Заголовки
                ws.append(['Cell', 'Value', 'Row', 'Col'])
                
                # Добавляем данные
                for cell_data in cells:
                    ws.append([
                        cell_data['cell'],
                        cell_data['value'],
                        cell_data['row'],
                        cell_data['col']
                    ])
        
        # Сохраняем результат
        result_wb.save(output_path)
    
    return categories


def print_color_summary(categories):
    """Вывод сводки по цветам"""
    print("Сводка по цветам ячеек:")
    print("="*50)
    
    for color, cells in categories.items():
        status_map = {
            'red': 'Не ответили на звонок',
            'yellow': 'Занесено в базу данных',
            'green': 'Занесено в базу данных',
            'gray': 'Только позвонили, но не занесли в базу данных',
            'none': 'Еще не сделано'
        }
        
        status = status_map.get(color, 'Неизвестная категория')
        print(f"{color.capitalize()}: {len(cells)} ячеек - {status}")
    
    print("="*50)


if __name__ == "__main__":
    # Пример использования
    file_path = "АСП_Многодетные.xlsx"
    output_path = "sorted_by_colors.xlsx"
    
    try:
        categories = sort_excel_by_colors(file_path, output_path)
        print_color_summary(categories)
        print(f"\nРезультаты сохранены в '{output_path}'")
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")