#!/usr/bin/env python3
"""Скрипт для сортировки семей в Excel файле по цветам заливки"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import PatternFill as OpenpyxlPatternFill
from collections import OrderedDict
import argparse
import sys
import os
from copy import copy


def get_cell_background_color(cell):
    """Получение цвета фона ячейки"""
    # Проверяем fill_type - если None или 'none', значит нет заливки
    if cell.fill.fill_type is None or cell.fill.fill_type == 'none':
        return None
    
    # Получаем цвета из разных возможных источников
    color_attrs = ['fgColor', 'bgColor', 'start_color', 'end_color']
    
    for attr_name in color_attrs:
        if hasattr(cell.fill, attr_name):
            color_obj = getattr(cell.fill, attr_name)
            if color_obj is not None and hasattr(color_obj, 'rgb') and color_obj.rgb is not None:
                rgb_color = color_obj.rgb
                # Проверяем, что это не "пустой" цвет (черный или прозрачный)
                if rgb_color not in ['00000000', '000000']:
                    return rgb_color
    
    return None


def group_families_by_color(file_path):
    """Группировка семей по цветам заливки"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    families_by_color = {}  # Словарь для хранения семей по цветам
    current_family_rows = []  # Текущая группа строк для семьи
    current_family_color = None  # Цвет текущей семьи
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        # Проверяем, есть ли значение в первом столбце (это начало новой семьи)
        first_cell = row[0]  # Первый столбец (A)
        
        if first_cell.value is not None and str(first_cell.value).strip() != "" and str(first_cell.value).isdigit():
            # Это начало новой семьи
            
            # Если у нас есть предыдущая семья, сохраним её
            if current_family_rows:
                if current_family_color not in families_by_color:
                    families_by_color[current_family_color] = []
                families_by_color[current_family_color].append(current_family_rows)
            
            # Начинаем новую семью
            current_family_rows = [row]
            current_family_color = get_cell_background_color(first_cell)
        else:
            # Это продолжение текущей семьи
            current_family_rows.append(row)
    
    # Не забываем добавить последнюю семью
    if current_family_rows:
        if current_family_color not in families_by_color:
            families_by_color[current_family_color] = []
        families_by_color[current_family_color].append(current_family_rows)
    
    return families_by_color


def create_sorted_excel(families_by_color, output_path):
    """Создание отсортированного Excel файла"""
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "Отсортировано по цветам"
    
    # Сортируем цвета для последовательного вывода (None в конец)
    sorted_colors = sorted(
        [c for c in families_by_color.keys() if c is not None], 
        key=lambda x: x if x is not None else ''
    )
    if None in families_by_color:
        sorted_colors.append(None)
    
    row_num = 1  # Номер строки в выходном файле
    
    for color in sorted_colors:
        if color not in families_by_color:
            continue
            
        families = families_by_color[color]
        
        # Добавляем заголовок с информацией о цвете
        if color:
            ws_output.cell(row=row_num, column=1, value=f"Цвет заливки: {color}")
            ws_output.cell(row=row_num, column=1).fill = OpenpyxlPatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            ws_output.cell(row=row_num, column=1, value="Без заливки")
        
        row_num += 1
        
        # Добавляем все семьи этого цвета
        for family in families:
            for row in family:
                for col_idx, cell in enumerate(row, 1):
                    ws_output.cell(row=row_num, column=col_idx, value=cell.value)
                    
                    # Копируем стиль ячейки
                    if cell.has_style:
                        ws_output.cell(row=row_num, column=col_idx).font = copy(cell.font)
                        ws_output.cell(row=row_num, column=col_idx).border = copy(cell.border)
                        ws_output.cell(row=row_num, column=col_idx).fill = copy(cell.fill)
                        ws_output.cell(row=row_num, column=col_idx).number_format = cell.number_format
                        ws_output.cell(row=row_num, column=col_idx).protection = copy(cell.protection)
                        ws_output.cell(row=row_num, column=col_idx).alignment = copy(cell.alignment)
                
                row_num += 1
            
            # Добавляем пустую строку между семьями
            row_num += 1
    
    # Сохраняем файл
    wb_output.save(output_path)


def print_color_statistics(families_by_color):
    """Вывод статистики по цветам"""
    print("Статистика по цветам заливки семей:")
    print("="*50)
    
    total_families = sum(len(families) for families in families_by_color.values())
    
    # Сортируем цвета для вывода
    sorted_items = sorted(
        [(k, v) for k, v in families_by_color.items() if k is not None],
        key=lambda x: x[0]
    )
    
    if None in families_by_color:
        sorted_items.append((None, families_by_color[None]))
    
    for color, families in sorted_items:
        count = len(families)
        if color:
            print(f"Цвет {color}: {count} семей ({count/total_families*100:.1f}%)")
        else:
            print(f"Без заливки: {count} семей ({count/total_families*100:.1f}%)")
    
    print(f"Всего семей: {total_families}")
    print("="*50)


def main():
    parser = argparse.ArgumentParser(description='Сортировка семей в Excel файле по цветам заливки')
    parser.add_argument('input_file', help='Путь к входному Excel файлу')
    parser.add_argument('-o', '--output', default='sorted_families_by_colors.xlsx',
                       help='Путь к выходному файлу (по умолчанию: sorted_families_by_colors.xlsx)')
    
    args = parser.parse_args()
    
    # Проверяем существование входного файла
    if not os.path.exists(args.input_file):
        print(f"Ошибка: файл {args.input_file} не найден")
        sys.exit(1)
    
    print(f"Обработка файла: {args.input_file}")
    
    try:
        # Группируем семьи по цветам
        families_by_color = group_families_by_color(args.input_file)
        
        # Выводим статистику
        print_color_statistics(families_by_color)
        
        # Создаем отсортированный Excel файл
        create_sorted_excel(families_by_color, args.output)
        
        print(f"Результаты сохранены в: {args.output}")
        
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()