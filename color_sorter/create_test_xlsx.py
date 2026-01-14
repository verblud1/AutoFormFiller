#!/usr/bin/env python3
"""Создание тестового Excel файла с цветной заливкой ячеек"""

import openpyxl
from openpyxl.styles import PatternFill


def create_test_file():
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Тестовый лист"
    
    # Определяем цвета заливки
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Красный
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Желтый
    green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # Зеленый
    gray_fill = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")  # Серый
    
    # Добавляем тестовые данные с разными цветами
    # Красные ячейки - не ответили на звонок
    ws['A1'] = "Не ответили на звонок"
    ws['A1'].fill = red_fill
    
    ws['A2'] = "Пропущенный звонок"
    ws['A2'].fill = red_fill
    
    # Желтые ячейки - занесено в базу данных
    ws['B1'] = "Занесено в базу данных"
    ws['B1'].fill = yellow_fill
    
    ws['B2'] = "В обработке"
    ws['B2'].fill = yellow_fill
    
    # Зеленые ячейки - занесено в базу данных
    ws['C1'] = "Обработано"
    ws['C1'].fill = green_fill
    
    ws['C2'] = "В базе"
    ws['C2'].fill = green_fill
    
    # Серые ячейки - только позвонили, но не занесли в базу данных
    ws['D1'] = "Позвонили"
    ws['D1'].fill = gray_fill
    
    ws['D2'] = "Без записи"
    ws['D2'].fill = gray_fill
    
    # Ячейки без заливки - еще не сделано
    ws['E1'] = "Не обработано"
    ws['E2'] = "Ожидает"
    
    # Добавим немного больше данных для тестирования
    for i in range(3, 6):
        # Красные
        ws.cell(row=i, column=1, value=f"Красный {i}")
        ws.cell(row=i, column=1).fill = red_fill
        
        # Желтые
        ws.cell(row=i, column=2, value=f"Желтый {i}")
        ws.cell(row=i, column=2).fill = yellow_fill
        
        # Зеленые
        ws.cell(row=i, column=3, value=f"Зеленый {i}")
        ws.cell(row=i, column=3).fill = green_fill
        
        # Серые
        ws.cell(row=i, column=4, value=f"Серый {i}")
        ws.cell(row=i, column=4).fill = gray_fill
        
        # Без цвета
        ws.cell(row=i, column=5, value=f"Без цвета {i}")
    
    # Сохраняем файл
    wb.save("test_colored_file.xlsx")
    print("Тестовый файл 'test_colored_file.xlsx' создан")


if __name__ == "__main__":
    create_test_file()